from rest_framework import viewsets, filters
from .models import Patient, Appointment
from .serializers import PatientSerializer, AppointmentSerializer
from rest_framework.decorators import api_view
from rest_framework.response import Response
from rest_framework import status
from django.shortcuts import render
import openpyxl
from django.db import transaction



@api_view(['GET', 'POST'])
def list_patients(request):
    if request.method == 'GET':
        patients = Patient.objects.all()
        serializer = PatientSerializer(patients, many=True)
        return Response(serializer.data)

    elif request.method == 'POST':
        serializer = PatientSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

def patients_view(request):
    return render(request, 'patients.html')




class PatientViewSet(viewsets.ModelViewSet):
    queryset = Patient.objects.all()
    serializer_class = PatientSerializer
    filter_backends = [filters.SearchFilter]
    search_fields = ['name', 'medical_history', 'date_of_birth', 'gender']

class AppointmentViewSet(viewsets.ModelViewSet):
    queryset = Appointment.objects.all()
    serializer_class = AppointmentSerializer
    filter_backends = [filters.SearchFilter]
    search_fields = ['patient__name', 'doctor_name', 'department', 'appointment_date']

@api_view(['POST'])
def upload_data(request):
    if 'file' not in request.FILES:
        return Response({'error': 'No file provided'}, status=status.HTTP_400_BAD_REQUEST)

    file = request.FILES['file']
    try:
        wb = openpyxl.load_workbook(file)
        patients_sheet = wb['Patients']
        appointments_sheet = wb['Appointments']

        with transaction.atomic():
            for row in patients_sheet.iter_rows(min_row=2, values_only=True):
                Patient.objects.create(
                    patient_id=row[0],
                    name=row[1],
                    contact_information=row[2],
                    medical_history=row[3],
                    date_of_birth=row[4],
                    gender=row[5]
                )
            
            for row in appointments_sheet.iter_rows(min_row=2, values_only=True):
                patient = Patient.objects.get(patient_id=row[1])
                Appointment.objects.create(
                    appointment_id=row[0],
                    patient=patient,
                    doctor_name=row[2],
                    department=row[3],
                    appointment_date=row[4],
                    appointment_time=row[5],
                    reason_for_visit=row[6]
                )

        return Response({'status': 'Data uploaded successfully'}, status=status.HTTP_201_CREATED)

    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)
